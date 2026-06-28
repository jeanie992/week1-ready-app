# -*- coding: utf-8 -*-
"""
benchmark.py : 경쟁사·피어 DART 자동수집 → 핵심 재무 벤치마킹 표

사용 : python benchmark.py "한무컨벤션" "GKL" "파라다이스" "조선호텔앤리조트" "호텔롯데" "호텔신라"
       (첫 회사 = focus, 강조) : 회사명은 DART 등록명 기준(예 GKL)
출력 : 재무마스터_자동.xlsx 가 있으면 '8.경쟁사벤치마킹' 탭 추가, 없으면 경쟁사벤치마킹.xlsx
주의 : 연결/별도·사업범위 차이(예 호텔신라=면세 포함) → 단순 비교 아님 : 회사별 기준연도 표기
"""
from __future__ import annotations
import sys, os, io, zipfile, json
from pathlib import Path
import os as _os; ROOT = Path(_os.environ.get("PROJECT_ROOT") or _os.getcwd())
sys.path.insert(0, str(Path(__file__).resolve().parent))
import dart_fetch as DF
import dart_parse as DP
import excel_style as S
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

PEER_DIR = ROOT / "02_analysis" / "_peers"
EOK = "#,##0.0;[Red]\\-#,##0.0;\\-;@"
PCT = "0.0%"
GROW = '+0.0%;[Red]-0.0%;"-"'


def _find(d, keys):
    nd = {e.replace(" ", ""): e for e in d}
    for k in keys:
        if k.replace(" ", "") in nd:
            return nd[k.replace(" ", "")]
    for k in keys:
        kk = k.replace(" ", "")
        for norm, orig in nd.items():
            if kk in norm:
                return orig
    return None


def _eok(v):
    return v / 1e8 if v is not None else None


def _amt(s):
    try:
        return float(str(s).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _api_metrics(key, corp, y4):
    """상장사 : DART 재무제표 API(정형 JSON) : 반환 (dict, 연도, 연결여부) 또는 (None,..)"""
    for year in (y4 - 1, y4 - 2):
        for fs in ("CFS", "OFS"):
            url = (f"{DF.BASE}/fnlttSinglAcntAll.json?crtfc_key={key}&corp_code={corp}"
                   f"&bsns_year={year}&reprt_code=11011&fs_div={fs}")
            try:
                j = json.loads(DF.http_get(url).decode("utf-8"))
            except Exception:
                continue
            lst = j.get("list") if j.get("status") == "000" else None
            if not lst:
                continue

            def g(sjs, ids, nms, col="thstrm_amount"):
                for it in lst:
                    if it.get("sj_div") not in sjs:
                        continue
                    aid = it.get("account_id", ""); anm = (it.get("account_nm") or "").replace(" ", "")
                    if aid in ids or any(x in anm for x in nms):
                        return _amt(it.get(col))
                return None
            rev = g(("IS", "CIS"), ("ifrs-full_Revenue", "dart_Revenue"), ("매출액", "영업수익", "수익(매출액)"))
            revp = g(("IS", "CIS"), ("ifrs-full_Revenue", "dart_Revenue"), ("매출액", "영업수익"), "frmtrm_amount")
            oi = g(("IS", "CIS"), ("dart_OperatingIncomeLoss", "ifrs-full_ProfitLossFromOperatingActivities"), ("영업이익",))
            ni = g(("IS", "CIS"), ("ifrs-full_ProfitLoss",), ("당기순이익",))
            asset = g(("BS",), ("ifrs-full_Assets",), ("자산총계",))
            debt = g(("BS",), ("ifrs-full_Liabilities",), ("부채총계",))
            eq = g(("BS",), ("ifrs-full_Equity",), ("자본총계",))
            if rev or asset:
                return {"매출": rev, "매출p": revp, "영업이익": oi, "순이익": ni,
                        "자산총계": asset, "부채총계": debt, "자본총계": eq}, year, (fs == "CFS")
    return None, None, None


def _doc_metrics(key, corp, name, y4):
    """비상장 : 감사보고서 문서 파싱"""
    PEER_DIR.mkdir(parents=True, exist_ok=True)
    fp = PEER_DIR / (name.replace("/", "_") + ".txt")
    if not fp.exists():
        reps = DF.list_reports(key, corp, f"{y4-2}0101", f"{y4}1231")
        ch = DF.pick_annual(reps, 1)
        if not ch:
            return None, None
        blob = DF.fetch_document(key, ch[0]["rcept_no"])
        texts = []
        with zipfile.ZipFile(io.BytesIO(blob)) as z:
            for n in z.namelist():
                if n.lower().endswith((".xml", ".html", ".htm", ".txt")):
                    texts.append(DF.xml_to_text(z.read(n)))
        fp.write_text("\n\n".join(texts), encoding="utf-8")
    m = DP.parse_report(fp)
    if not m or not m["years"]:
        return None, None
    cur = m["years"][0]; prev = m["years"][1] if len(m["years"]) > 1 else None

    def v(stmt, keys, y):
        e = _find(m[stmt], keys)
        return m[stmt].get(e, {}).get(y) if e else None
    return {"매출": v("IS", ["매출", "영업수익", "매출액"], cur),
            "매출p": v("IS", ["매출", "영업수익", "매출액"], prev) if prev else None,
            "영업이익": v("IS", ["영업이익"], cur), "순이익": v("IS", ["당기순이익"], cur),
            "자산총계": v("BS", ["자산총계"], cur), "부채총계": v("BS", ["부채총계"], cur),
            "자본총계": v("BS", ["자본총계"], cur)}, cur


def peer_metrics(key, name):
    corp = DF.get_corp_code(key, name)
    end = "".join(filter(str.isdigit, os.environ.get("TODAY", "")))[:8] or "20991231"
    y4 = int(end[:4]) if end[:4].isdigit() else 2026
    d, year, cfs = _api_metrics(key, corp, y4)        # 상장사 우선
    src = f"API({'연결' if cfs else '별도'})"
    if d is None:                                      # 비상장 폴백
        d, year = _doc_metrics(key, corp, name, y4)
        src = "감사보고서"
        if d is None:
            return None
    rev, revp = d["매출"], d["매출p"]
    return {"회사": name, "연도": year, "매출": rev,
            "매출YoY": (rev / revp - 1) if (rev and revp and revp > 0) else None,
            "영업이익": d["영업이익"], "영업이익률": (d["영업이익"] / rev) if (d["영업이익"] is not None and rev) else None,
            "순이익": d["순이익"], "자산총계": d["자산총계"],
            "부채비율": (d["부채총계"] / d["자본총계"]) if (d["부채총계"] is not None and d["자본총계"]) else None,
            "출처": src}


def build_tab(rows, focus):
    headers = ["회사", "연도", "매출", "매출 YoY", "영업이익", "영업이익률", "순이익", "자산총계", "부채비율"]
    mp = ROOT / "03_output" / "재무마스터_자동.xlsx"
    if mp.exists():
        wb = load_workbook(mp); target = mp; tname = "8.경쟁사벤치마킹"
    else:
        wb = Workbook(); wb.remove(wb.active)
        target = ROOT / "03_output" / "경쟁사벤치마킹.xlsx"; tname = "경쟁사벤치마킹"
    if tname in wb.sheetnames:
        del wb[tname]
    ws = wb.create_sheet(tname)
    S.label(ws.cell(2, 2), "경쟁사·피어 벤치마킹 (DART 최신 보고서)", bold=True)
    ws.cell(3, 2).value = S.clean_text("단위 억원·% : 연결/별도·사업범위 차이 주의(예 호텔신라=면세 포함) : 회사별 기준연도 표기")
    ws.cell(3, 2).font = S.FONT_NOTE
    hr = 4
    for j, h in enumerate(headers):
        S.header(ws.cell(hr, 2 + j), h)
    r = hr + 1
    for d in rows:
        isF = (d["회사"] == focus)
        S.label(ws.cell(r, 2), d["회사"], bold=isF)
        S.num(ws.cell(r, 3), d["연도"], "0")
        S.num(ws.cell(r, 4), _eok(d["매출"]), EOK, bold=isF)
        S.num(ws.cell(r, 5), d["매출YoY"], GROW)
        S.num(ws.cell(r, 6), _eok(d["영업이익"]), EOK)
        S.num(ws.cell(r, 7), d["영업이익률"], PCT)
        S.num(ws.cell(r, 8), _eok(d["순이익"]), EOK)
        S.num(ws.cell(r, 9), _eok(d["자산총계"]), EOK)
        S.num(ws.cell(r, 10), d["부채비율"], PCT)
        if isF:
            for cc in range(2, 11):
                ws.cell(r, cc).fill = S.FILL_BAND
        r += 1
    ws.column_dimensions["A"].width = S.W_MARGIN
    ws.column_dimensions["B"].width = 18
    for j in range(1, len(headers)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 11
    ws.sheet_view.showGridLines = False
    wb._sheets.sort(key=lambda s: 1 if s.title.startswith("원본_") else 0)
    wb.save(target)
    return target, tname


def main():
    key = DF.load_key()
    companies = [a for a in sys.argv[1:] if not a.startswith("-")]
    if not companies:
        companies = ["한무컨벤션", "GKL", "파라다이스", "조선호텔앤리조트", "호텔롯데", "호텔신라"]
    focus = companies[0]
    rows = []
    for name in companies:
        try:
            r = peer_metrics(key, name)
            if r:
                rows.append(r)
                print(f"  OK {name} ({r['연도']}, {r.get('출처','')}) 매출 {(_eok(r['매출']) or 0):,.0f}억 영업이익률 {((r['영업이익률'] or 0)*100):.1f}%")
            else:
                print(f"  skip {name} : 보고서/파싱 실패")
        except Exception as e:
            print(f"  fail {name} : {e}")
    if not rows:
        sys.exit("[오류] 수집된 회사가 없습니다")
    target, tname = build_tab(rows, focus)
    print(f"저장 : {target} ({tname})")


if __name__ == "__main__":
    main()
