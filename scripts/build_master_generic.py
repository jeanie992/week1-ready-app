# -*- coding: utf-8 -*-
"""
build_master_generic.py : 어떤 회사든 DART + 내부 Excel 로 '원본참조 수식형' 재무 마스터 생성

구조 (BCG 모델 방식) :
  - 분석 탭(앞) : 값이 아니라 **원본 시트를 참조하는 수식**(=원본!셀/1e8) : 원본 바뀌면 자동 갱신·추적 가능
  - 원본 탭(뒤) : DART 파싱 원본(원 단위) + 내부 Excel 원본 그대로
탭 : Cover / 1.손익 / 2.재무상태 / 3.현금흐름 / 4.사업부손익 / 5.비용구조_매출대비 / 원본_손익·재무상태·현금흐름·내부
인사이트 : ANTHROPIC_API_KEY 있으면 Claude, 없으면 로컬 Claude Code(Max), 없으면 규칙 기반
사용 : python build_master_generic.py "회사명"
문체 : 끝 온점 없음, 하이픈 대신 콜론
"""
from __future__ import annotations
import sys, re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os as _os; ROOT = Path(_os.environ.get("PROJECT_ROOT") or _os.getcwd())

import excel_style as S
import dart_parse
import llm_notes

OUT = ROOT / "03_output" / "재무마스터_자동.xlsx"
F = 100000000          # 원 → 억원
EOK = "#,##0.0;[Red]\\-#,##0.0;\\-;@"
PCT = "0.0%"
GROW = '+0.0%;[Red]-0.0%;"-"'
PPF = '+0.0"%p";[Red]-0.0"%p";"-"'
RAW = "#,##0"
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

SUBTOTAL_KW = ("총계", "영업수익", "영업비용", "영업이익", "당기순이익", "현금흐름",
               "순이익", "법인세비용차감전", "포괄손익", "유동자산", "비유동자산",
               "유동부채", "비유동부채", "자본금", "이익잉여금")


def is_sub(label):
    n = re.sub(r"\s", "", str(label))
    return any(k.replace(" ", "") in n for k in SUBTOTAL_KW)


def _val(tidy, ent, cat, y):
    s = tidy[(tidy["entity"] == ent) & (tidy["category"] == cat) & (tidy["year"] == y)]["value"]
    return float(s.sum()) if len(s) else None


def _find(ents, keys):
    for k in keys:
        for e in ents:
            if str(e).strip() == k:
                return e
    for k in keys:
        for e in ents:
            if k in str(e):
                return e
    return None


def _ctx(tidy, iyears, cats, rev, oi):
    """NOTE 인과추적용 내부 맥락 : 사업부별 매출·영업이익 최근 추세를 LLM 에 전달"""
    if tidy is None or not iyears:
        return ""
    ys = iyears[-3:] if len(iyears) >= 3 else iyears
    lines = ["[내부 사업부 손익 (억원) : 매출 변화의 출처 추적용]"]
    for cat in cats:
        rv = " ".join(f"{y}:{(_val(tidy, rev, cat, y) or 0)/1e8:.0f}" for y in ys)
        ov = " ".join(f"{(_val(tidy, oi, cat, y) or 0)/1e8:.0f}" for y in ys)
        lines.append(f"{cat} 매출 {rv} / 영업이익 {ov}")
    return "\n".join(lines)


def _numfmt(cell, fmt, bold=False):
    cell.font = S.FONT_NUM_B if bold else S.FONT_NUM
    cell.alignment = S.AL_NUM
    cell.number_format = fmt
    cell.border = S.BORDER_THIN
    return cell


def _note(cell, text):
    cell.value = S.clean_text(text) if text else None
    cell.font = S.FONT_NOTE
    cell.alignment = WRAP
    cell.border = S.BORDER_THIN


# ── 원본 시트 (뒤) ──────────────────────────────────────────────────────────
def raw_statement(wb, key, years, order, data):
    ws = wb.create_sheet(f"원본_{key}")
    S.header(ws.cell(1, 1), "계정")
    for j, y in enumerate(years):
        S.header(ws.cell(1, 2 + j), str(y))
    ref, r = {}, 2
    for acc in order:
        if all(data[acc].get(y) is None for y in years):
            continue
        ws.cell(r, 1).value = acc; ws.cell(r, 1).font = S.FONT_LABEL; ws.cell(r, 1).border = S.BORDER_THIN
        for j, y in enumerate(years):
            v = data[acc].get(y)
            if v is not None:
                c = ws.cell(r, 2 + j); c.value = v; _numfmt(c, RAW)
                ref[(acc, y)] = f"'{ws.title}'!{c.coordinate}"
        r += 1
    ws.column_dimensions["A"].width = 26
    for j in range(len(years)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 16
    ws.sheet_view.showGridLines = False
    return ref


def raw_internal(wb, tidy, years):
    ws = wb.create_sheet("원본_내부")
    S.header(ws.cell(1, 1), "구분"); S.header(ws.cell(1, 2), "계정")
    for j, y in enumerate(years):
        S.header(ws.cell(1, 3 + j), str(y))
    ref, r = {}, 2
    for cat, ent in tidy[["category", "entity"]].drop_duplicates().itertuples(index=False):
        ws.cell(r, 1).value = str(cat); ws.cell(r, 1).font = S.FONT_LABEL; ws.cell(r, 1).border = S.BORDER_THIN
        ws.cell(r, 2).value = str(ent); ws.cell(r, 2).font = S.FONT_LABEL; ws.cell(r, 2).border = S.BORDER_THIN
        for j, y in enumerate(years):
            v = _val(tidy, ent, cat, y)
            if v is not None:
                c = ws.cell(r, 3 + j); c.value = v; _numfmt(c, RAW)
                ref[(cat, ent, y)] = f"'원본_내부'!{c.coordinate}"
        r += 1
    ws.column_dimensions["A"].width = 14; ws.column_dimensions["B"].width = 18
    for j in range(len(years)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 14
    ws.sheet_view.showGridLines = False
    return ref


# ── 분석 시트 (앞, 원본참조 수식) ───────────────────────────────────────────
def stmt_tab(wb, name, title, years, order, data, notes, ref):
    ws = wb.create_sheet(name)
    S.label(ws.cell(2, 2), title, bold=True)
    ws.cell(3, 2).value = S.clean_text("단위 : 억원 : 값은 뒤 원본시트 참조 수식 : 보조 YoY·CAGR")
    ws.cell(3, 2).font = S.FONT_NOTE
    hr = 4
    heads = ["항목"] + [str(y) for y in years] + [f"YoY '{str(years[-1])[2:]}",
             f"CAGR '{str(years[0])[2:]}-'{str(years[-1])[2:]}", "NOTE"]
    for j, h in enumerate(heads):
        S.header(ws.cell(hr, 2 + j), h)
    ny = len(years); vc = [get_column_letter(3 + j) for j in range(ny)]
    r = hr + 1
    for acc in order:
        if all((acc, y) not in ref for y in years):
            continue
        sub = is_sub(acc)
        S.label(ws.cell(r, 2), acc, bold=sub)
        for j, y in enumerate(years):
            c = ws.cell(r, 3 + j)
            if (acc, y) in ref:
                c.value = f"={ref[(acc, y)]}/{F}"; _numfmt(c, EOK, bold=sub)
            else:
                _numfmt(c, EOK)
        yoy = ws.cell(r, 3 + ny)
        if ny >= 2:
            yoy.value = f'=IFERROR(({vc[-1]}{r}/{vc[-2]}{r})-1,"")'
        _numfmt(yoy, GROW)
        cagr = ws.cell(r, 3 + ny + 1)
        cagr.value = f'=IFERROR(({vc[-1]}{r}/{vc[0]}{r})^(1/{ny-1})-1,"")'
        _numfmt(cagr, GROW)
        _note(ws.cell(r, 3 + ny + 2), notes.get(acc, ""))
        if sub:
            for cc in range(2, 3 + ny + 2):
                ws.cell(r, cc).fill = S.FILL_BAND
        r += 1
    ws.column_dimensions["A"].width = S.W_MARGIN; ws.column_dimensions["B"].width = 28
    for j in range(ny + 2):
        ws.column_dimensions[get_column_letter(3 + j)].width = 11
    ws.column_dimensions[get_column_letter(3 + ny + 2)].width = 50
    ws.sheet_view.showGridLines = False


def seg_tab(wb, years, cats, rev, oi, totcat, ref2):
    ws = wb.create_sheet("4.사업부손익")
    S.label(ws.cell(2, 2), f"사업부별 {rev}·{oi}·이익률 ({years[0]}~{years[-1]})", bold=True)
    ws.cell(3, 2).value = S.clean_text("단위 : 억원 (이익률 %) : 값은 원본_내부 참조 수식 : 보조 CAGR·YoY")
    ws.cell(3, 2).font = S.FONT_NOTE
    hr = 4
    S.header(ws.cell(hr, 2), "사업부 / 지표")
    for j, y in enumerate(years):
        S.header(ws.cell(hr, 3 + j), str(y))
    base = 3 + len(years)
    S.header(ws.cell(hr, base), "CAGR"); S.header(ws.cell(hr, base + 1), "YoY")
    ny = len(years); vc = [get_column_letter(3 + j) for j in range(ny)]
    order = [c for c in cats if c != totcat] + ([totcat] if totcat in cats else [])
    r = hr + 1
    for cat in order:
        isT = cat == totcat
        mrow = r
        for ent, lm in [(rev, "매출"), (oi, "영업이익")]:
            S.label(ws.cell(r, 2), f"{cat} : {lm}", bold=isT)
            for j, y in enumerate(years):
                c = ws.cell(r, 3 + j)
                if (cat, ent, y) in ref2:
                    c.value = f"={ref2[(cat, ent, y)]}/{F}"; _numfmt(c, EOK, bold=isT)
                else:
                    _numfmt(c, EOK)
            ws.cell(r, base).value = f'=IFERROR(({vc[-1]}{r}/{vc[0]}{r})^(1/{ny-1})-1,"")'; _numfmt(ws.cell(r, base), GROW)
            ws.cell(r, base + 1).value = f'=IFERROR(({vc[-1]}{r}/{vc[-2]}{r})-1,"")'; _numfmt(ws.cell(r, base + 1), GROW)
            if isT:
                for cc in range(2, base + 2):
                    ws.cell(r, cc).fill = S.FILL_BAND
            r += 1
        # 영업이익률 = 영업이익행 / 매출행 (앞 셀 참조)
        S.label(ws.cell(r, 2), f"{cat} : 영업이익률", bold=isT)
        for j in range(ny):
            col = vc[j]
            ws.cell(r, 3 + j).value = f'=IFERROR({col}{mrow+1}/{col}{mrow},"")'; _numfmt(ws.cell(r, 3 + j), PCT)
        ws.cell(r, base).value = f"=({vc[-1]}{r}-{vc[0]}{r})*100"; _numfmt(ws.cell(r, base), PPF)
        ws.cell(r, base + 1).value = f"=({vc[-1]}{r}-{vc[-2]}{r})*100"; _numfmt(ws.cell(r, base + 1), PPF)
        if isT:
            for cc in range(2, base + 2):
                ws.cell(r, cc).fill = S.FILL_BAND
        r += 1
    ws.column_dimensions["A"].width = S.W_MARGIN; ws.column_dimensions["B"].width = 20
    for j in range(ny + 2):
        ws.column_dimensions[get_column_letter(3 + j)].width = 9
    ws.sheet_view.showGridLines = False


def cost_tab(wb, tidy, years, totcat, rev, ref2):
    ws = wb.create_sheet("5.비용구조_매출대비")
    S.label(ws.cell(2, 2), f"공통비율 : 매출 100원당 항목 비중 ({totcat})", bold=True)
    ws.cell(3, 2).value = S.clean_text("단위 : % : =원본_내부 항목/매출 수식 : 보조 Δ(첫→끝, %p)")
    ws.cell(3, 2).font = S.FONT_NOTE
    show = years if len(years) <= 6 else [years[0], years[len(years) // 2], years[-3], years[-2], years[-1]]
    hr = 4
    S.header(ws.cell(hr, 2), "항목")
    for j, y in enumerate(show):
        S.header(ws.cell(hr, 3 + j), str(y))
    S.header(ws.cell(hr, 3 + len(show)), "Δ %p")
    ns = len(show); vc = [get_column_letter(3 + j) for j in range(ns)]
    ents = list(tidy[tidy["category"] == totcat]["entity"].unique())
    ents.sort(key=lambda e: abs(_val(tidy, e, totcat, years[-1]) or 0), reverse=True)
    r = hr + 1
    for e in ents:
        S.label(ws.cell(r, 2), e, bold=(e == rev))
        for j, y in enumerate(show):
            c = ws.cell(r, 3 + j)
            if (totcat, e, y) in ref2 and (totcat, rev, y) in ref2:
                c.value = f'=IFERROR({ref2[(totcat, e, y)]}/{ref2[(totcat, rev, y)]},"")'; _numfmt(c, PCT)
            else:
                _numfmt(c, PCT)
        ws.cell(r, 3 + ns).value = f"=({vc[-1]}{r}-{vc[0]}{r})*100"; _numfmt(ws.cell(r, 3 + ns), PPF)
        if e == rev:
            for cc in range(2, 4 + ns):
                ws.cell(r, cc).fill = S.FILL_BAND
        r += 1
    ws.column_dimensions["A"].width = S.W_MARGIN; ws.column_dimensions["B"].width = 22
    for j in range(ns + 1):
        ws.column_dimensions[get_column_letter(3 + j)].width = 10
    ws.sheet_view.showGridLines = False


def cover(wb, company, dart_years, internal_years, used_llm):
    ws = wb.create_sheet("Cover")
    S.header(ws.cell(2, 2), f"{company} 재무 마스터 (원본참조 수식형)")
    meta = [("대상회사", company),
            ("DART 기간", f"{dart_years[0]}~{dart_years[-1]}" if dart_years else "없음"),
            ("내부 P&L 기간", f"{internal_years[0]}~{internal_years[-1]}" if internal_years else "없음"),
            ("구조", "분석 탭(앞) = 원본 탭(뒤) 참조 수식 : 원본 바뀌면 자동 갱신·추적 가능"),
            ("단위", "분석 억원 / 원본 원 (÷1e8)"),
            ("인사이트", "Claude/Max" if used_llm else "규칙 기반"),
            ("탭", "1~3 DART : 4~5 내부 P&L : 원본_* (뒤)"),
            ("문체", "문장 끝 온점 없음 : 하이픈 대신 콜론")]
    r = 4
    for k, v in meta:
        S.label(ws.cell(r, 2), k, bold=True)
        ws.cell(r, 3).value = S.clean_text(v); ws.cell(r, 3).font = S.FONT_LABEL; ws.cell(r, 3).alignment = WRAP
        r += 1
    ws.column_dimensions["A"].width = S.W_MARGIN; ws.column_dimensions["B"].width = 14; ws.column_dimensions["C"].width = 80
    ws.sheet_view.showGridLines = False


def load_internal():
    try:
        import analyze_excel as AX
        for fn, sh, raw in AX.load_tables():
            t, note, warns = AX.to_tidy(raw)
            if t is None or t.empty:
                continue
            cats = [c for c in t["category"].unique() if str(c).strip()]
            ents = list(t["entity"].unique())
            rev = _find(ents, ["매출", "영업수익", "매출액", "수익"])
            oi = _find(ents, ["영업이익"])
            if rev and oi and len(cats) >= 2:
                tot = next((c for c in cats if str(c).strip() in ("Total", "합계", "계", "총계")), cats[0])
                return t, sorted(int(y) for y in t["year"].unique()), cats, rev, oi, tot
    except Exception as e:
        print(f"[내부] 분석 불가 : {e}")
    return None, [], [], None, None, None


def main():
    company = next((a for a in sys.argv[1:] if not a.startswith("-")), "회사")
    use_llm = "--no-llm" not in sys.argv
    m = dart_parse.parse_all()
    dyears = m["years"]
    tidy, iyears, cats, rev, oi, totcat = load_internal()

    wb = Workbook(); wb.remove(wb.active)
    notes_all = {}
    cover(wb, company, dyears, iyears, use_llm)

    # 분석 탭 (앞) — 원본참조
    if dyears:
        notes_all = llm_notes.generate(m, dyears, company=company,
                                       extra_context=_ctx(tidy, iyears, cats, rev, oi), use_llm=use_llm)
        nt = {st: {a: n for (s2, a), n in notes_all.items() if s2 == st} for st in ("IS", "BS", "CF")}
        rIS = raw_statement(wb, "손익", dyears, m["order"]["IS"], m["IS"])
        rBS = raw_statement(wb, "재무상태", dyears, m["order"]["BS"], m["BS"])
        rCF = raw_statement(wb, "현금흐름", dyears, m["order"]["CF"], m["CF"])
        stmt_tab(wb, "1.손익계산서", f"{company} 손익계산서 (억원)", dyears, m["order"]["IS"], m["IS"], nt["IS"], rIS)
        stmt_tab(wb, "2.재무상태표", f"{company} 재무상태표 (억원)", dyears, m["order"]["BS"], m["BS"], nt["BS"], rBS)
        stmt_tab(wb, "3.현금흐름표", f"{company} 현금흐름표 (억원)", dyears, m["order"]["CF"], m["CF"], nt["CF"], rCF)
    if tidy is not None:
        r2 = raw_internal(wb, tidy, iyears)
        seg_tab(wb, iyears, cats, rev, oi, totcat, r2)
        cost_tab(wb, tidy, iyears, totcat, rev, r2)

    if len(wb.sheetnames) <= 1:
        sys.exit("[오류] 분석할 자료가 없습니다 (DART 또는 내부 Excel 필요)")

    # 원본 탭을 맨 뒤로 정렬
    front = ["Cover", "1.손익계산서", "2.재무상태표", "3.현금흐름표", "4.사업부손익", "5.비용구조_매출대비"]
    wb._sheets.sort(key=lambda s: (1, s.title) if s.title.startswith("원본_") else (0, front.index(s.title) if s.title in front else 99))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("저장 :", OUT)
    print(f"탭 : {wb.sheetnames}")
    print(f"  DART {dyears or '없음'} : 내부 {iyears or '없음'} : 노트 {len(notes_all)}건")


if __name__ == "__main__":
    main()
