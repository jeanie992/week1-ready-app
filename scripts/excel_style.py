# -*- coding: utf-8 -*-
"""
excel_style.py : 재무 모델 양식을 따르는 재사용 스타일 모듈

목적 : 생성하는 엑셀의 폰트·색·테두리·숫자서식을 일관되게 적용
팔레트 : 그린 테마(헤더 진그린 + 흰 글자, 음수 빨강) : 숫자 Calibri, 한글 맑은 고딕

문체 규칙(중요) :
  1) 문장 끝에 온점(.) 금지
  2) 하이픈(-) 으로 잇지 말 것 → 콜론( : ) 사용 (공백+콜론+공백)
  3) 라벨은 간결한 명사구, 단위는 별도 열에 표기 ( 억 원 / % / 원 )
"""
from __future__ import annotations
import re
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# ── 컬러 팔레트 (테마에서 추출, openpyxl 은 ARGB 'FF'+HEX) ──────────────────
C = {
    "text":      "FF000000",  # dk1 기본 텍스트
    "white":     "FFFFFFFF",  # lt1 배경/헤더 글자
    "gray_dk":   "FF575757",  # dk2 보조 텍스트
    "gray_band": "FFF2F2F2",  # lt2 구분 밴드
    "green":     "FF29BA74",  # accent1 메인 그린
    "green_dk":  "FF197A56",  # accent2 헤더 배경(진한 그린)
    "lime":      "FFD4D533",  # accent3
    "teal":      "FF3EAD92",  # accent4
    "gray":      "FF6E6F73",  # accent5
    "navy":      "FF295E7E",  # accent6
    "input":     "FFFFFFCC",  # 입력/가정 셀(연노랑)
    "pos_green": "FF00B050",  # 강조 텍스트 그린
    "blue":      "FF0070C0",  # 강조 텍스트 블루
    "red":       "FFFF0000",  # 음수/경고
}

# ── 폰트 (숫자/영문 = Calibri 11, 한글 라벨/헤더 = 맑은 고딕) ────────────────
FONT_NUM      = Font(name="Calibri", size=11, color=C["text"])
FONT_NUM_B    = Font(name="Calibri", size=11, bold=True, color=C["text"])
FONT_LABEL    = Font(name="맑은 고딕", size=11, color=C["text"])
FONT_LABEL_B  = Font(name="맑은 고딕", size=11, bold=True, color=C["text"])
FONT_HEADER   = Font(name="맑은 고딕", size=10, bold=True, color=C["white"])   # 진한 그린 배경 위
FONT_WARN     = Font(name="맑은 고딕", size=11, color=C["red"])                # 음수/경고
FONT_NOTE     = Font(name="맑은 고딕", size=10, color=C["gray_dk"])            # 주석

# ── 채움 ────────────────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor=C["green_dk"])   # 헤더행
FILL_BAND   = PatternFill("solid", fgColor=C["gray_band"])  # 소계/구분 밴드
FILL_INPUT  = PatternFill("solid", fgColor=C["input"])      # 입력/가정
FILL_NONE   = PatternFill(fill_type=None)

# ── 테두리 ──────────────────────────────────────────────────────────────────
_thin = Side(style="thin", color="FFBFBFBF")
_med  = Side(style="medium", color="FF000000")
_hair = Side(style="hair", color="FFD9D9D9")
BORDER_THIN     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_TOTAL    = Border(left=_thin, right=_thin, top=_thin, bottom=_med)   # 합계행 강조
BORDER_BOTTOM   = Border(bottom=_thin)

# ── 정렬 ────────────────────────────────────────────────────────────────────
AL_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_LABEL  = Alignment(horizontal="left",  vertical="center")
AL_NUM    = Alignment(horizontal="right", vertical="center")
AL_CENTER = Alignment(horizontal="center", vertical="center")

# ── 숫자 서식 ────────────────────────────────────────────────────────────────
NUMFMT = {
    "int":      "#,##0",
    "int_red":  "#,##0;[Red]\\-#,##0;\\-;@",   # 음수 빨강, 0 은 대시
    "pct":      "0%",
    "pct1":     "0.0%",
    "won":      "#,##0",
}

# ── 권장 열너비 ──────────────────────────────────────────────────────────────
W_MARGIN = 2.6     # A열 여백/들여쓰기
W_LABEL  = 25.6    # 라벨열
W_DATA   = 15.6    # 데이터열
W_UNIT   = 6.0     # 단위열


# ── 문체 정리기 ──────────────────────────────────────────────────────────────
def clean_text(s):
    """문체 규칙 적용 : 끝 온점 제거, ' - ' → ' : ' (날짜/숫자 범위의 붙은 하이픈은 보존)"""
    if not isinstance(s, str):
        return s
    s = s.strip()
    s = re.sub(r"\s[-–—]\s", " : ", s)   # 공백 하이픈 공백 → 콜론
    s = re.sub(r"[.]+$", "", s)            # 끝 온점 제거
    return s


# ── 셀 스타일 적용 헬퍼 ───────────────────────────────────────────────────────
def header(cell, text=None):
    if text is not None:
        cell.value = clean_text(text)
    cell.font, cell.fill, cell.alignment, cell.border = FONT_HEADER, FILL_HEADER, AL_HEADER, BORDER_THIN
    return cell


def label(cell, text=None, bold=False, indent=0):
    if text is not None:
        cell.value = clean_text(text)
    cell.font = FONT_LABEL_B if bold else FONT_LABEL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    cell.border = BORDER_THIN
    return cell


def num(cell, value=None, fmt="int_red", bold=False):
    if value is not None:
        cell.value = value
    cell.font = FONT_NUM_B if bold else FONT_NUM
    cell.alignment = AL_NUM
    cell.number_format = NUMFMT.get(fmt, fmt)
    cell.border = BORDER_THIN
    return cell


def band(cell):
    cell.fill = FILL_BAND
    return cell


def total_row(cells):
    for c in cells:
        c.font = FONT_NUM_B if c.font.name == "Calibri" else FONT_LABEL_B
        c.fill = FILL_BAND
        c.border = BORDER_TOTAL
    return cells


def input_cell(cell, value=None, fmt="int"):
    if value is not None:
        cell.value = value
    cell.fill = FILL_INPUT
    cell.font = FONT_NUM
    cell.number_format = NUMFMT.get(fmt, fmt)
    cell.alignment = AL_NUM
    cell.border = BORDER_THIN
    return cell


def set_widths(ws, label_col="B", data_cols=None, label_w=W_LABEL, data_w=W_DATA):
    """A열 여백 + 라벨열 + 데이터열 너비 일괄 지정"""
    ws.column_dimensions["A"].width = W_MARGIN
    ws.column_dimensions[label_col].width = label_w
    for col in (data_cols or []):
        ws.column_dimensions[col].width = data_w


def write_table(ws, top, left, headers, rows, units=None, fmt="int_red"):
    """간단한 스타일 표 작성
    headers : 1행 헤더 리스트(첫 칸은 라벨 헤더)
    rows    : [(라벨, [값...]) , ...]
    units   : 라벨 옆 단위열 텍스트(없으면 생략)
    반환 : 마지막으로 쓴 행 번호
    """
    from openpyxl.utils import get_column_letter
    r = top
    for j, h in enumerate(headers):
        header(ws.cell(r, left + j), h)
    r += 1
    for lbl, vals in rows:
        label(ws.cell(r, left), lbl)
        for j, v in enumerate(vals):
            cell = ws.cell(r, left + 1 + j)
            if v is None or v == "":
                num(cell, None, fmt)
            else:
                num(cell, v, fmt)
        r += 1
    # 데이터 열너비
    cols = [get_column_letter(left)] + [get_column_letter(left + 1 + j) for j in range(len(headers) - 1)]
    ws.column_dimensions["A"].width = W_MARGIN
    ws.column_dimensions[cols[0]].width = W_LABEL
    for c in cols[1:]:
        ws.column_dimensions[c].width = W_DATA
    return r - 1


def append_tab(path, title, headers, rows, subtitle="", keep_raw_last=True):
    """기존 워크북(path)에 스타일 탭을 추가(같은 이름 있으면 교체)
    Claude 가 운영지표·요약 등 판단형 탭을 붙일 때 사용 : 스크립트 자동탭 위에 깊이를 더한다
    headers : 컬럼 헤더 리스트(첫 칸 = 라벨 헤더) : "NOTE" 포함 가능
    rows    : [(cells, fmt), ...] : cells = 헤더 길이에 맞춘 값(문자=라벨, 숫자=수치) : fmt = "num"/"pct"/"eok"
    예) append_tab("03_output/재무마스터_자동.xlsx", "6.운영지표",
                   ["지표","단위","OPCC","OPI","NOTE"],
                   [(["영업이익률","%",0.356,0.141,"OPCC 최고"], "pct"),
                    (["ADR","천원",330,212,"인상여력"], "num")])
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    fmts = {"num": "#,##0", "pct": "0.0%", "pct1": "0.0%",
            "eok": "#,##0.0;[Red]\\-#,##0.0;\\-;@", "won": "#,##0"}
    wb = load_workbook(path)
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    label(ws.cell(2, 2), title, bold=True)
    if subtitle:
        ws.cell(3, 2).value = clean_text(subtitle); ws.cell(3, 2).font = FONT_NOTE
    hr = 4
    for j, h in enumerate(headers):
        header(ws.cell(hr, 2 + j), h)
    r = hr + 1
    for cells, fmt in rows:
        for j, v in enumerate(cells):
            cell = ws.cell(r, 2 + j)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                num(cell, v, fmts.get(fmt, fmt))
            elif j < len(headers) and str(headers[j]).upper() == "NOTE":
                cell.value = clean_text(v) if v else None
                cell.font = FONT_NOTE
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = BORDER_THIN
            else:
                label(cell, "" if v is None else str(v))
        r += 1
    ws.column_dimensions["A"].width = W_MARGIN
    ws.column_dimensions["B"].width = 22
    for j in range(1, len(headers)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 40 if str(headers[j]).upper() == "NOTE" else 11
    ws.sheet_view.showGridLines = False
    if keep_raw_last:        # 원본_ 시트는 항상 맨 뒤로
        wb._sheets.sort(key=lambda s: 1 if s.title.startswith("원본_") else 0)
    wb.save(path)
    return title


if __name__ == "__main__":
    # 데모 : 모듈이 정상 적용되는지 작은 표 생성
    from openpyxl import Workbook
    from pathlib import Path
    wb = Workbook(); ws = wb.active; ws.title = "demo"
    ws["B2"] = clean_text("사업부별 매출 - 영업이익 (2025년 기준).")  # 규칙 적용 확인용
    ws["B2"].font = FONT_LABEL_B
    last = write_table(
        ws, top=4, left=2,
        headers=["사업부", "매출", "영업이익", "영업이익률"],
        rows=[("사업부 A", [100, 20, None]),
              ("사업부 B", [80, 10, None]),
              ("사업부 C", [50, 15, None]),
              ("Total", [230, 45, None])],
    )
    ws.cell(4, 5).value = "억 원"
    out = Path(__file__).resolve().parent.parent / "02_analysis" / "_style_demo.xlsx"
    wb.save(out)
    print("데모 저장 :", out)
    print("B2 정리 결과 :", ws["B2"].value)
